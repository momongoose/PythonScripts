# import module
import openpyxl
import csv

# load excel with its path
wrkbk = openpyxl.load_workbook(r"path")

sh = wrkbk.active
IDs = []
# iterate through excel and display data
for i in range(1, sh.max_row + 1):
    for j in range(1, sh.max_column + 1):
        cell_obj = sh.cell(row=i, column=j)
        if str(cell_obj.value) == "None":
            continue
        IDs.append(str(cell_obj.value))
print(IDs)
with open('csv.csv', 'w') as file:
    for i in IDs:
        file.write("text,?,?,?,?,0,?,number,?,no,no,no,no,#{},?,time,?,?,Available,no,no,?,name,datetime,?,?,?,?,no,?,?,?,?,?,?\n".format(i))