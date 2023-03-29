from openpyxl import load_workbook

fn2 = r'path'

workbook2 = load_workbook(filename=fn2)

sheet2 = workbook2.active
szvnArray = []
nfcArray = []
x = 0
for i in range(sheet2.max_row):
    f = i + 1
    if f > 1:
        name = str(sheet2.cell(f, 3).value).replace(" ","")
        szvnArray.append(name)
        nfcArray.append(str(sheet2.cell(f, 4).value))

workbook2.save(filename="Anwesenheit.xlsx")

fn = r'path'

workbook = load_workbook(filename=fn)
szvnSheet = []
sheet = workbook.active
x = 0
for i in range(sheet.max_row):
    f = i + 1
    if f > 7:
        b = 0
        for szvn in szvnArray:
            if str(sheet.cell(f, 2).value) == szvn:
                x += 1
                pos = str("D{}".format(str(f)))
                sheet[pos] = nfcArray[b]
            b += 1
        szvnSheet.append(str(sheet.cell(f, 2).value))

for szvn in szvnArray:
    if szvn in szvnSheet:
        continue
    else:
        print(szvn)
print(x)
workbook.save(filename="Check.xlsx")
